#!/usr/bin/env python3
"""
04_enrich_excel.py — Verrijk het Excel-bestand met kleuren en logo-bestandsnamen.

Wat doet dit script:
- Leest het bestaande hockeyclubs Excel-bestand
- Voegt kolommen toe: kleur_1, kleur_2, kleur_1_naam, kleur_2_naam, logo_bestand, kleur_bron
- Matcht op clubnaam (fuzzy) met de kleurdata
- Respecteert handmatige overrides (kleur_bron = "handmatig" wordt NIET overschreven)

Gebruik:
    python scripts/04_enrich_excel.py
    python scripts/04_enrich_excel.py --input pad/naar/bestand.xlsx
    python scripts/04_enrich_excel.py --force   # overschrijf ook handmatige overrides

Input:
    data/club_kleuren.json
    data/club_logo_mapping.json
    data/hockeyclubs_nederland.xlsx (of via --input)

Output:
    data/hockeyclubs_nederland_verrijkt.xlsx
"""

import json
import sys
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

# Standaard input Excel (kan overschreven worden met --input)
DEFAULT_INPUT = DATA_DIR / "hockeyclubs_nederland.xlsx"


def slugify(name: str) -> str:
    """Maak een simpele slug van een clubnaam voor matching."""
    name = name.lower().strip()
    # Verwijder haakjes en inhoud
    name = re.sub(r'\([^)]*\)', '', name)
    # Verwijder speciale tekens
    name = re.sub(r'[^a-z0-9\s]', '', name)
    name = re.sub(r'\s+', '-', name.strip())
    return name


def find_best_match(clubnaam: str, slug_map: dict) -> str | None:
    """
    Probeer een clubnaam te matchen met een slug uit de mapping.
    Probeert meerdere strategieën.
    """
    club_slug = slugify(clubnaam)
    
    # Exacte match
    if club_slug in slug_map:
        return club_slug
    
    # Bevat-match (slug bevat clubnaam of andersom)
    for slug in slug_map:
        if club_slug in slug or slug in club_slug:
            return slug
    
    # Eerste woord match
    first_word = club_slug.split('-')[0] if '-' in club_slug else club_slug
    if len(first_word) > 3:
        for slug in slug_map:
            if slug.startswith(first_word) or first_word in slug:
                return slug
    
    return None


def main():
    force = "--force" in sys.argv
    
    # Bepaal input bestand
    input_path = DEFAULT_INPUT
    for i, arg in enumerate(sys.argv):
        if arg == "--input" and i + 1 < len(sys.argv):
            input_path = Path(sys.argv[i + 1])
    
    if not input_path.exists():
        print(f"FOUT: Excel bestand niet gevonden: {input_path}")
        print(f"Kopieer je hockeyclubs Excel naar {DEFAULT_INPUT}")
        print(f"Of gebruik: python scripts/04_enrich_excel.py --input pad/naar/bestand.xlsx")
        sys.exit(1)
    
    # Laad kleurdata
    kleuren_path = DATA_DIR / "club_kleuren.json"
    mapping_path = DATA_DIR / "club_logo_mapping.json"
    
    if not kleuren_path.exists():
        print("FOUT: data/club_kleuren.json niet gevonden. Draai eerst stap 03.")
        sys.exit(1)
    
    with open(kleuren_path, "r", encoding="utf-8") as f:
        kleuren = json.load(f)
    
    with open(mapping_path, "r", encoding="utf-8") as f:
        logo_mapping = {club["slug"]: club for club in json.load(f)}
    
    # Bouw reverse mapping: slug → kleuren
    slug_map = {slug: True for slug in kleuren}
    
    # Laad Excel
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Zoek bestaande kolommen
    headers = [cell.value for cell in ws[1]]
    print(f"Bestaande kolommen: {headers}")
    
    # Bepaal of kleur-kolommen al bestaan
    new_cols = ["kleur_1", "kleur_2", "kleur_1_naam", "kleur_2_naam", "logo_bestand", "kleur_bron"]
    existing_col_indices = {}
    for col_name in new_cols:
        if col_name in headers:
            existing_col_indices[col_name] = headers.index(col_name) + 1  # 1-indexed
    
    # Voeg nieuwe kolommen toe als ze nog niet bestaan
    next_col = len(headers) + 1
    for col_name in new_cols:
        if col_name not in existing_col_indices:
            existing_col_indices[col_name] = next_col
            ws.cell(row=1, column=next_col, value=col_name)
            ws.cell(row=1, column=next_col).font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            ws.cell(row=1, column=next_col).fill = PatternFill('solid', fgColor='2E7D32')
            next_col += 1
    
    # Zoek de clubnaam kolom
    clubnaam_col = None
    for i, h in enumerate(headers):
        if h and "clubnaam" in h.lower():
            clubnaam_col = i + 1
            break
    
    if not clubnaam_col:
        print("FOUT: Kolom 'Clubnaam' niet gevonden in Excel.")
        sys.exit(1)
    
    # Verwerk elke rij
    matched = 0
    skipped_manual = 0
    not_found = 0
    
    for row in range(2, ws.max_row + 1):
        clubnaam = ws.cell(row=row, column=clubnaam_col).value
        if not clubnaam:
            continue
        
        # Check of dit een handmatige override is
        bron_col = existing_col_indices["kleur_bron"]
        current_bron = ws.cell(row=row, column=bron_col).value
        if current_bron == "handmatig" and not force:
            skipped_manual += 1
            continue
        
        # Zoek match
        best_slug = find_best_match(clubnaam, {**slug_map, **{s: True for s in logo_mapping}})
        
        if best_slug and best_slug in kleuren:
            colors = kleuren[best_slug]
            ws.cell(row=row, column=existing_col_indices["kleur_1"]).value = colors["kleur_1"]
            ws.cell(row=row, column=existing_col_indices["kleur_2"]).value = colors["kleur_2"]
            ws.cell(row=row, column=existing_col_indices["kleur_1_naam"]).value = colors.get("kleur_1_naam", "")
            ws.cell(row=row, column=existing_col_indices["kleur_2_naam"]).value = colors.get("kleur_2_naam", "")
            ws.cell(row=row, column=existing_col_indices["logo_bestand"]).value = f"{best_slug}.jpg"
            ws.cell(row=row, column=existing_col_indices["kleur_bron"]).value = "logo_analyse"
            
            # Kleurblokjes als achtergrondkleur (voor visuele check in Excel)
            try:
                hex1 = colors["kleur_1"].lstrip("#")
                ws.cell(row=row, column=existing_col_indices["kleur_1"]).fill = PatternFill('solid', fgColor=hex1)
                # Witte of zwarte tekst afhankelijk van achtergrond
                brightness = (int(hex1[0:2], 16) * 299 + int(hex1[2:4], 16) * 587 + int(hex1[4:6], 16) * 114) / 1000
                text_color = "000000" if brightness > 128 else "FFFFFF"
                ws.cell(row=row, column=existing_col_indices["kleur_1"]).font = Font(name='Arial', size=10, color=text_color)
                
                hex2 = colors["kleur_2"].lstrip("#")
                ws.cell(row=row, column=existing_col_indices["kleur_2"]).fill = PatternFill('solid', fgColor=hex2)
                brightness2 = (int(hex2[0:2], 16) * 299 + int(hex2[2:4], 16) * 587 + int(hex2[4:6], 16) * 114) / 1000
                text_color2 = "000000" if brightness2 > 128 else "FFFFFF"
                ws.cell(row=row, column=existing_col_indices["kleur_2"]).font = Font(name='Arial', size=10, color=text_color2)
            except Exception:
                pass
            
            matched += 1
        else:
            not_found += 1
            if not_found <= 10:
                print(f"  Geen match: '{clubnaam}'")
    
    # Kolom breedtes
    ws.column_dimensions[openpyxl.utils.get_column_letter(existing_col_indices["kleur_1"])].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(existing_col_indices["kleur_2"])].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(existing_col_indices["kleur_1_naam"])].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(existing_col_indices["kleur_2_naam"])].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(existing_col_indices["logo_bestand"])].width = 30
    ws.column_dimensions[openpyxl.utils.get_column_letter(existing_col_indices["kleur_bron"])].width = 15
    
    # Opslaan
    output_path = DATA_DIR / "hockeyclubs_nederland_verrijkt.xlsx"
    wb.save(output_path)
    
    print(f"\nKlaar!")
    print(f"  Gematched: {matched}")
    print(f"  Handmatig (overgeslagen): {skipped_manual}")
    print(f"  Geen match gevonden: {not_found}")
    print(f"\nOpgeslagen: {output_path}")


if __name__ == "__main__":
    main()
